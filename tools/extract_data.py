import os
import openpyxl
import pymysql
import datetime
import sys

sys.path.append('../')
import time
from tools import dingding
dingsender = dingding.dingding()
from etc import SendMail
from etc import MysqlPool


class Etract(object):
    def __init__(self):
        self.dba_pool = MysqlPool.Mysql('dba')
        self.file_name = 'data.xlsx'

    def get_query_config(self):
        sql_get_config = 'select id,`sql`,`host`,`db`,mail_to from auto_extract_data where `label`=1  order by id limit 1 ;'
        result1 = self.dba_pool.getAll(sql_get_config)
        if result1:
            sql2 = 'update auto_extract_data set label=0 where id={0};'.format(result1[0]['id'])
            self.dba_pool.update(sql2)
        return result1

    def get_query_result(self, sql, ip, db):

        
            sql1 = u'''{0}'''.format(sql)
            conn = pymysql.connect(
                host=ip,
                user='root',
                passwd='Dba@2019huohua!',
                port=3306,
                database=db,
                charset='utf8mb4',
                cursorclass=pymysql.cursors.DictCursor
            )
            cur = conn.cursor()  # 创建游标
            cur.execute(sql1)  # 执行sql命令
            result2 = cur.fetchall()  # 获取执行的返回结果
            # print(result)
            cur.close()
            conn.close()  # 关闭mysql 连接
            return (result2)
        




    def generate_table(self, result2):

        
            """
            生成excel表格
            :return:
            """
            # 删除已存在的文件
            if os.path.exists(self.file_name):
                os.remove(self.file_name)

            result = result2

            # print(result)
            if not result:
                print("查询结果为空")
                return False

            # 创建excel对象
            f = openpyxl.Workbook()
            sheet1 = f.create_sheet('Sheet1', 0)

            # 第一行结果
            row0 = result[0]
            # 列字段
            column_names = list(row0)

            # 写第一行，也就是列所在的行
            for i in range(0, len(row0)):
                sheet1.cell(row=1, column=i + 1, value=column_names[i])

            # 写入多行
            # 行坐标，从第2行开始，也是1
            for row_id in range(1, len(result) + 1):
                # 列坐标
                for col_id in range(len(column_names)):
                    # 写入的值
                    value = result[row_id - 1][column_names[col_id]]
                    # 判断为日期时
                    if isinstance(value, datetime.datetime):
                        value = result[row_id - 1][column_names[col_id]].strftime('%Y-%m-%d %H:%M:%S')

                    # 写入表格
                    sheet1.cell(row=row_id + 1, column=col_id + 1, value=value)

            # 保存文件
            f.save(self.file_name)

            # 判断文件是否存在
            if not os.path.exists(self.file_name):
                print("生成excel失败")
                return False

            print("生成excel成功")
            return True

      



if __name__ == '__main__':
    while 1:
        try:
            extract = Etract()
            result1 = extract.get_query_config()
            if result1:
                result2 = extract.get_query_result(result1[0]['sql'], result1[0]['host'], result1[0]['db'])
                extract.generate_table(result2)
                mail1 = SendMail.SendMail()
                receiver = eval(result1[0]['mail_to'])
                mail1.Send(receiver, '提取数据', time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()), 'data.xlsx')
                os.remove('data.xlsx')
                time.sleep(5)

            else:
                time.sleep(5)
        except Exception as e:
            dingsender.send_msg('提取数据报错，请check参数，具体错误：{0} \n'.format(e))
            continue

